# Features of Timesheet.py
# Create a timesheets.xlxs ONCE, with 12 sheets with 12 month names
# which means it needs to check if the file exists before creating it
# takes command line strings : start, lunch start,lunch end,end and record the time when each command is applied
# when end is issued ,calculate total working time that day start-end-lunch break
# at the end of the month calculate total working hours
#TODO How to handle holidays?



from datetime import datetime
from dateutil.relativedelta import *
import calendar
from openpyxl import Workbook, load_workbook
import os.path
import click



# This method returns the current time
def get_timenow():
    now = datetime.now()
    return now


# this function returns todays date
def get_todays_date():

    return datetime.today().strftime("%d-%m-%Y")


# open timesheet.py, go sheet named today.month , check if todays date is is in there , if not make an entry in the next row
#TODO check if todays date is already in the file if yes delete that row and make new row for today
def enter_todays_date_in_file():
    wb = load_workbook(filename="timesheet.xlsx")
    today = get_timenow()
    month = today.strftime("%B")
    ws = wb[month]
    print(ws)
    print(ws.max_row)

    ws.cell(column=1, row=ws.max_row + 1, value=get_todays_date())
    wb.save(filename="timesheet.xlsx")


# this method creates a new timesheet if it does not exist in the root path
def create_timesheet():
    #TODO Month summary page
    # check if a timesheet.xlsx exists
    if os.path.isfile("timesheet.xlsx"):
        print("File exists")
    else:
        worksheet = []  # declare a list
        now = get_timenow()
        workbook = Workbook()

        for i in range(0, 12):
            worksheet.append(workbook.create_sheet(calendar.month_name[i + 1], i))

        for sheet in workbook:
            sheet["A1"] = "Date"
            sheet["B1"] = "Day"
            sheet["C1"] = "Start"
            sheet["D1"] = "Lunch Start"
            sheet["E1"] = "Lunch End"
            sheet["F1"] = "End"
            sheet["G1"] = "Total time"
            sheet["H1"] = "Comment"

        workbook.save(filename="timesheet.xlsx")
    # if not then create it with sheets with month names and row headers "Date" "Day" , "Start Time" ,"Lunch Start","Lunch End","End Time"


def enter_start_time():
    enter_todays_date_in_file() 
    #TODO get current time and write to the row with todays date and at cell start time
    return


def enter_end_time():
    #TODO Get current time and add it to End time column
    #TODO Calculate total working time 
    #TODO Add up total month working time
    return


@click.command()
@click.argument("Entry")
def main(entry):
    entry = entry
    print(entry)

    if entry == "start":
        enter_start_time()
    elif entry == "end":
        enter_end_time()
    elif entry == "create timesheet":
        create_timesheet()
    else:
        print("Invalid Entry")

    print(get_todays_date())
    


main()
